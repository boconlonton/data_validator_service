import os
import io
import logging

from contextlib import ExitStack
from collections import namedtuple

import boto3

from openpyxl import load_workbook

from aurora_data_api import connect

from validator.loader import get_key_from_event
from validator.loader import get_info_from_key
from validator.loader import get_data_from_s3
from validator.loader import get_existed_emails
from validator.loader import get_headers

from validator.writer import ExcelWriter
from validator.writer import DBWriter
from validator.writer import TaskStatus

from validator.engine import Validator

from validator.engine.errors import MissingHeadersError
from validator.engine.errors import TaskError
from validator.engine.errors import TaskMessage

logger = logging.getLogger()
logger.setLevel(logging.INFO)

client = boto3.client('s3')

S3_BUCKET = os.environ.get('S3_BUCKET')

Validator._PROVINCES = get_data_from_s3(client,
                                        bucket=S3_BUCKET,
                                        key='metadata/province.json')
Validator._DISTRICTS = get_data_from_s3(client,
                                        bucket=S3_BUCKET,
                                        key='metadata/district.json')
Validator._WARDS = get_data_from_s3(client,
                                    bucket=S3_BUCKET,
                                    key='metadata/ward.json')
Validator._WORKING_STATUS = get_data_from_s3(client,
                                             bucket=S3_BUCKET,
                                             key='metadata/working_status.json')
Validator._USER_TYPE = get_data_from_s3(client,
                                        bucket=S3_BUCKET,
                                        key='metadata/user_type.json')
Validator._CONTRACT = get_data_from_s3(client,
                                       bucket=S3_BUCKET,
                                       key='metadata/contract.json')
Validator._ETHNICS = get_data_from_s3(client,
                                      bucket=S3_BUCKET,
                                      key='metadata/ethnic.json')
Validator._RELIGION = get_data_from_s3(client,
                                       bucket=S3_BUCKET,
                                       key='metadata/religion.json')
Validator._TEACHING_TITLE = get_data_from_s3(client,
                                             bucket=S3_BUCKET,
                                             key='metadata/teaching_title.json')
Validator._ACADEMIC_TITLE = get_data_from_s3(client,
                                             bucket=S3_BUCKET,
                                             key='metadata/academic_title.json')
Validator._DEGREE = get_data_from_s3(client,
                                     bucket=S3_BUCKET,
                                     key='metadata/degree.json')
Validator._EMAILS = get_existed_emails(instance_arn=os.getenv('INSTANCE_ARN'),
                                       secret_arn=os.getenv('SECRET_ARN'),
                                       db_name=os.getenv('DB_NAME'))


def executioner(*, original_key, task, file_obj):
    response = client.get_object(
        Bucket=S3_BUCKET,
        Key=original_key,
    )
    body = response.get('Body')
    wb = load_workbook(io.BytesIO(body.read()))
    ws = wb['Data']
    if not ws:
        task.update_task(msg=TaskMessage.MISSING_DATA_SHEET,
                         status=TaskStatus.COMPLETED)
        logger.warning(f'MISSING_DATA_SHEET')
        return
    headers = get_headers(worksheet=ws)
    try:
        Validator.validate_headers(headers=headers)
    except MissingHeadersError as e:
        task.update_task(msg=TaskMessage.MISSING_DATA_SHEET,
                         status=TaskStatus.COMPLETED,
                         details=str(e))
        logger.warning(f'MISSING_HEADERS: {e}')
        return
    builder = namedtuple('Builder', field_names=headers)
    for row in ws.iter_rows(min_row=2,
                            max_col=ws.max_column,
                            max_row=ws.max_row,
                            values_only=True):
        obj = Validator(builder(*row))
        obj.validate()
    Validator.check_duplicate()

    if Validator.bad_stream:
        error_file = ExcelWriter(headers=headers,
                                 stream=Validator.bad_stream,
                                 obj_key=file_obj)
        error_file.upload_to_s3(client=client, bucket_name=S3_BUCKET)

    if Validator.duplicate_stream:
        duplicate_stream = list(Validator.duplicate_stream)
        Validator.stats['duplicated'] = len(duplicate_stream)
        duplicated_file = ExcelWriter(headers=headers,
                                      stream=duplicate_stream,
                                      obj_key=file_obj,
                                      duplicate_file=True)
        duplicated_file.upload_to_s3(client=client, bucket_name=S3_BUCKET)
    task.write_users(headers=headers, stream=Validator.valid_stream)
    total = Validator.stats.get('total')
    failed = Validator.stats.get('failed')
    duplicated = Validator.stats.get('duplicated')
    task.update_task(status=TaskStatus.COMPLETED,
                     total=total,
                     failed=failed,
                     duplicated=duplicated)
    logger.info(f'COMPLETED: Task(id={task.task_id},'
                f'total={total},failed={failed},duplicated={duplicated})')
    return


def lambda_handler(event, *args, **kwargs):
    original_key = get_key_from_event(event=event)
    try:
        file_obj = get_info_from_key(key=original_key)
    except TaskError as e:
        logger.warning(f'FILE_NAME_ERROR: {e}')
        return
    with ExitStack() as stack:
        conn = stack.enter_context(
            connect(
                os.getenv('INSTANCE_ARN'),
                os.getenv('SECRET_ARN'),
                database=os.getenv('DB_NAME')
            )
        )
        cursor = stack.enter_context(conn.cursor())
        task = DBWriter(task_id=file_obj.task_id, cursor=cursor)
        try:
            executioner(original_key=original_key,
                        task=task,
                        file_obj=file_obj)
        except Exception as e:
            logger.error('ENGINE_ERROR', exc_info=e)
            task.update_task(msg=TaskMessage.ENGINE_ERROR,
                             status=TaskStatus.FAILED)
    return
