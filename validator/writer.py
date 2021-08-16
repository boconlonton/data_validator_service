from enum import Enum

from typing import List
from typing import Any

from tempfile import NamedTemporaryFile

from openpyxl import Workbook

from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill

from validator.loader import file_nt
from validator.engine.errors import TaskMessage

error_style = NamedStyle(name='error')
error_style.fill = PatternFill(start_color='00FF0000',
                               end_color='00FF0000',
                               fill_type='solid')


class TaskStatus(Enum):
    IN_PROGRESS = 0
    COMPLETED = 1
    FAILED = 2


class ExcelWriter:

    def __init__(self,
                 *,
                 headers: List[str],
                 stream: List[Any],
                 obj_key: file_nt,
                 duplicate_file: bool = False):
        self.headers = headers
        self.obj_key = obj_key
        self.duplicated_file = duplicate_file
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Data'
        self.ws.append([
            " ".join(field.split('_')).capitalize()
            for field in self.headers
        ])
        self.write_sheet(stream)

    @property
    def duplicated_file(self):
        return self._duplicated

    @duplicated_file.setter
    def duplicated_file(self, value):
        self._duplicated = value
        if not value:
            self.headers.append('error_details')

    def write_sheet(self, stream):
        for row_idx, item in enumerate(stream):
            self.write_rows_excel(row_idx=row_idx + 2, data=item)

    def write_rows_excel(self, *, row_idx, data: dict):
        """Write a row to the Worksheet

        Args:
            data (dict): specify row data
            row_idx (int): specify row index
        """
        col_idx = 1
        temp = data.get('data')._asdict()
        errors = data.pop('errors', None)
        for name, value in temp.items():
            cell = self.ws.cell(column=col_idx, row=row_idx, value=value)
            if errors and name in errors:
                cell.style = error_style
            col_idx += 1
        self.ws.cell(column=col_idx,
                     row=row_idx,
                     value=data.get('error_msg'))

    def upload_to_s3(self, *, client, bucket_name):
        if self.duplicated_file:
            post_fix = 'duplicates'
        else:
            post_fix = 'errors'
        obj_key = (f'output/{self.obj_key.task_id}'
                   f'/{self.obj_key.file.name}_{post_fix}.xlsx')
        with NamedTemporaryFile() as tmp:
            self.wb.save(tmp.name)
            tmp.seek(0)
            client.upload_fileobj(tmp,
                                  bucket_name,
                                  obj_key)


class DBWriter:

    def __init__(self,
                 *,
                 task_id: str,
                 cursor):
        self._cursor = cursor
        self.task_id = task_id

    @property
    def task_id(self):
        return self.task_id

    @task_id.setter
    def task_id(self, value):
        """Create a task in DB"""
        query = 'INSERT INTO "public"."task" (task_id) VALUES (:task_id)'
        self._cursor.execute(query, {'task_id': int(value)})

    def update_task(self,
                    *,
                    status: TaskStatus,
                    msg: TaskMessage = None,
                    details: str = None,
                    total: int = 0,
                    failed: int = 0,
                    duplicated: int = 0):
        """Update details of a task

        Args:
            status (str): specify task status
            msg (TaskMessage): specify error message
            details (str): specify details of error
            total (int): specify total records
            failed (int): specify total failed records
            duplicated (int): specify total duplicated records
        """
        query = (
            'UPDATE "public"."task" SET '
            'status = :status,'
            'message = :msg,'
            'details = :details,'
            'total = :total,'
            'failed = :failed,'
            'duplicated = :duplicated '
            'WHERE task_id = :task_id'
        )
        self._cursor.execute(
            query,
            {
                'status': status,
                'msg': msg,
                'details': details,
                'total': total,
                'failed': failed,
                'duplicated': duplicated,
                'task_id': self.task_id
            }
        )

    def write_users(self, *, headers, stream):
        pass
